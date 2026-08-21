from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


SOURCE_PATH = Path(r"C:\Users\User\Documents\Islamic Front\tmp_Anjuman_IFB_accounts.xlsx")
OUTPUT_PATH = Path(r"C:\Users\User\OneDrive\Documents 1\Anjuman_IFB_accounts_simple_v6.xlsx")

ENTRY_END_ROW = 5104
SETUP_CATEGORY_START = 5

COLORS = {
    "green": "0F5132",
    "green_soft": "E8F3EC",
    "gold_soft": "F6EFCF",
    "cream": "F8F5ED",
    "border": "CBD8D0",
    "red_soft": "FCE8E6",
    "white": "FFFFFF",
}

NORMALIZE_MAP = {
    "APSP CAMP MASJID WORK": "APSP Camp Masjid Work",
    "APSP CAMP WORK": "APSP Camp Masjid Work",
    "CC CAMERA  OFFICE AND OUTSIDE": "CCTV Cameras (Office and Outside)",
    "FREE MEDICAL CAMP EXPENSE": "Free Medical Camp Expenses",
    "MUNCIPAL TAX": "Municipal Tax",
    "muncipal taxes": "Municipal Tax",
    "R BLOCK CONSTRUCTION": "R Block Construction",
    "SHABEBARATH": "Shab-e-Barat Expenses",
    "miladunnabi expenses": "Milad-un-Nabi Expenses",
    "marhum khana expenses": "Marhum Khana Expenses",
}


def set_named_range(workbook, name: str, attr_text: str) -> None:
    workbook.defined_names.pop(name, None)
    workbook.defined_names.add(DefinedName(name, attr_text=attr_text))


def border() -> Border:
    return Border(
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        bottom=Side(style="thin", color=COLORS["border"]),
    )


def style_cell(cell, fill: str | None = None, bold: bool = False, color: str = "000000", align: str = "left") -> None:
    cell.border = border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.font = Font(bold=bold, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def ensure_setup_sheet(workbook):
    if "CATEGORY_SETUP" in workbook.sheetnames:
        old_sheet = workbook["CATEGORY_SETUP"]
        workbook.remove(old_sheet)
    if "ADD_CATEGORY" in workbook.sheetnames:
        sheet = workbook["ADD_CATEGORY"]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet("ADD_CATEGORY")
    return sheet


def ensure_helper_sheet(workbook):
    if "_SYSTEM" in workbook.sheetnames:
        sheet = workbook["_SYSTEM"]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet("_SYSTEM")
    return sheet


def ensure_readme_sheet(workbook):
    if "README" in workbook.sheetnames:
        sheet = workbook["README"]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet("README")
    return sheet


def clear_data_validations(sheet) -> None:
    sheet.data_validations.dataValidation = []


def main() -> None:
    workbook = load_workbook(SOURCE_PATH)
    entry = workbook["ENTRY"]
    summary = workbook["SUMMARY"]
    details = workbook["DETAILS"]

    normalized_counts = Counter()
    raw_aliases: dict[str, set[str]] = {}

    for row in range(5, ENTRY_END_ROW + 1):
        value = entry[f"D{row}"].value
        if value in (None, ""):
            continue
        raw_value = str(value).strip()
        clean_value = NORMALIZE_MAP.get(raw_value, raw_value)
        raw_aliases.setdefault(clean_value, set()).add(raw_value)
        if raw_value in NORMALIZE_MAP:
            entry[f"D{row}"] = clean_value
            normalized_counts[clean_value] += 1

    years = sorted(
        {
            int(entry[f"B{row}"].value)
            for row in range(5, ENTRY_END_ROW + 1)
            if entry[f"B{row}"].value not in (None, "")
        }
    )

    ordered_categories: list[str] = []
    seen: set[str] = set()

    for row in range(SETUP_CATEGORY_START, 300):
        value = entry[f"M{row}"].value
        if value in (None, ""):
            continue
        clean = NORMALIZE_MAP.get(str(value), str(value)).strip()
        if clean not in seen:
            ordered_categories.append(clean)
            seen.add(clean)

    for row in range(5, ENTRY_END_ROW + 1):
        value = entry[f"D{row}"].value
        if value in (None, ""):
            continue
        clean = str(value).strip()
        if clean not in seen:
            ordered_categories.append(clean)
            seen.add(clean)

    for row in range(SETUP_CATEGORY_START, 300):
        entry[f"M{row}"] = None

    for idx, category in enumerate(ordered_categories, start=SETUP_CATEGORY_START):
        entry[f"M{idx}"] = category

    entry.delete_cols(11, 10)

    setup = ensure_setup_sheet(workbook)
    readme = ensure_readme_sheet(workbook)
    setup["A1"] = "Add New Category Here"
    setup["A2"] = "Type a new category only in column A below this list. If you rename an existing category here, reports and filters will use the new name."
    setup["A4"] = "Category Name"

    style_cell(setup["A1"], COLORS["green"], bold=True, color=COLORS["white"])
    style_cell(setup["A2"], COLORS["cream"])
    style_cell(setup["A4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")

    setup.row_dimensions[1].height = 24
    setup.row_dimensions[2].height = 36

    setup["B4"] = "Old Name"
    setup["C4"] = "Filter List"
    style_cell(setup["B4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    style_cell(setup["C4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    setup["C5"] = "All"
    style_cell(setup["C5"], COLORS["gold_soft"])

    for idx, category in enumerate(ordered_categories, start=SETUP_CATEGORY_START):
        setup[f"A{idx}"] = category
        style_cell(setup[f"A{idx}"], COLORS["green_soft"])
        setup[f"B{idx}"] = category
        style_cell(setup[f"B{idx}"], COLORS["cream"])
        setup[f"C{idx + 1}"] = category
        style_cell(setup[f"C{idx + 1}"], COLORS["cream"])

    setup.column_dimensions["A"].width = 34
    setup.column_dimensions["B"].hidden = True
    setup.column_dimensions["C"].hidden = True
    for col in ["D", "E", "F", "G"]:
        setup.column_dimensions[col].hidden = True

    helper = ensure_helper_sheet(workbook)
    helper["A4"] = "Months"
    helper["B4"] = "Types"
    helper["C4"] = "Years"
    helper["D4"] = "Type Filter"
    helper["E4"] = "Year Filter"
    helper["F4"] = "Month Filter"
    helper["G4"] = "Current Category"
    helper["H4"] = "Alias"
    for idx, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=5):
        helper[f"A{idx}"] = month
        helper[f"F{idx + 1}"] = month
    helper["B5"] = "Receipt"
    helper["B6"] = "Expense"
    helper["D5"] = "All"
    helper["D6"] = "Receipt"
    helper["D7"] = "Expense"
    helper["E5"] = "All"
    helper["F5"] = "All"
    for idx, year in enumerate(years, start=5):
        helper[f"C{idx}"] = year
        helper[f"E{idx + 1}"] = year

    alias_row = 5
    for idx, category in enumerate(ordered_categories, start=SETUP_CATEGORY_START):
        helper[f"G{alias_row}"] = f'=ADD_CATEGORY!$A${idx}'
        helper[f"H{alias_row}"] = category
        alias_row += 1
        for alias in sorted(raw_aliases.get(category, set())):
            if alias == category:
                continue
            helper[f"G{alias_row}"] = f'=ADD_CATEGORY!$A${idx}'
            helper[f"H{alias_row}"] = alias
            alias_row += 1
    helper.sheet_state = "hidden"

    readme["A1"] = "ముందు ఇది చదవండి"
    readme["A2"] = "ఈ ఫైల్ వాడే విధానం"
    readme["A4"] = "1. ENTRY షీట్‌లో మాత్రమే వివరాలు నమోదు చేయండి."
    readme["A5"] = "2. నెల, సంవత్సరం, రకం, కేటగిరీ, మొత్తం మాత్రమే నమోదు చేయండి."
    readme["A6"] = "3. కొత్త కేటగిరీ కావాలంటే ADD_CATEGORY షీట్‌లో ఒకసారి మాత్రమే జోడించండి."
    readme["A7"] = "4. తర్వాత ENTRY లో column D dropdown నుండి ఆ కేటగిరీని ఎంచుకోండి."
    readme["A8"] = "5. SUMMARY మరియు DETAILS షీట్లు ఆటోమేటిక్‌గా update అవుతాయి."
    readme["A9"] = "6. Islamic Front Board ఈ రికార్డులు 04-Aug-2023 నుండి నిర్వహిస్తోంది."
    readme["A10"] = "Short English:"
    readme["A11"] = "Use ENTRY for data entry. Use ADD_CATEGORY only to add or rename a category once."
    readme["A12"] = "Islamic Front Board has handled these records from August 4, 2023."
    style_cell(readme["A1"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    style_cell(readme["A2"], COLORS["gold_soft"], bold=True)
    for ref in ["A4", "A5", "A6", "A7", "A8", "A9", "A10", "A11", "A12"]:
        style_cell(readme[ref], COLORS["cream"])
    readme.column_dimensions["A"].width = 90
    readme.row_dimensions[1].height = 26
    readme.row_dimensions[2].height = 24
    for row in [4, 5, 6, 7, 8, 9, 11, 12]:
        readme.row_dimensions[row].height = 24

    set_named_range(workbook, "EntryMonthList", "_SYSTEM!$A$5:$A$16")
    set_named_range(workbook, "EntryTypeList", "_SYSTEM!$B$5:$B$6")
    set_named_range(workbook, "EntryYearList", '_SYSTEM!$C$5:INDEX(_SYSTEM!$C:$C,MATCH(9.99999999999999E+307,_SYSTEM!$C:$C))')
    set_named_range(workbook, "EntryCategoryList", 'ADD_CATEGORY!$A$5:INDEX(ADD_CATEGORY!$A:$A,MATCH("zzz",ADD_CATEGORY!$A:$A))')
    set_named_range(workbook, "FilterTypeList", "_SYSTEM!$D$5:$D$7")
    set_named_range(workbook, "FilterCategoryList", 'ADD_CATEGORY!$C$5:INDEX(ADD_CATEGORY!$C:$C,MATCH("zzz",ADD_CATEGORY!$C:$C))')
    set_named_range(workbook, "FilterYearList", '_SYSTEM!$E$5:INDEX(_SYSTEM!$E:$E,MATCH(9.99999999999999E+307,_SYSTEM!$E:$E))')
    set_named_range(workbook, "FilterMonthList", "_SYSTEM!$F$5:$F$17")

    clear_data_validations(entry)
    clear_data_validations(summary)
    clear_data_validations(details)

    month_formula_range = "_SYSTEM!$A$5:$A$16"
    for row in range(5, ENTRY_END_ROW + 1):
        entry[f"G{row}"] = '=IF(OR(A{0}="",B{0}=""),"",DATE(B{0},MATCH(A{0},{1},0),1))'.format(row, month_formula_range)

    dv_month = DataValidation(type="list", formula1="=EntryMonthList", allow_blank=True)
    dv_year = DataValidation(type="list", formula1="=EntryYearList", allow_blank=True)
    dv_type = DataValidation(type="list", formula1="=EntryTypeList", allow_blank=True)
    dv_category = DataValidation(type="list", formula1="=EntryCategoryList", allow_blank=True)
    dv_category.errorTitle = "Category not allowed"
    dv_category.error = "Do not type a new spelling here. Add it once in ADD_CATEGORY and then choose it from the list."
    dv_category.showErrorMessage = True
    dv_amount = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=True)

    for validation in [dv_month, dv_year, dv_type, dv_category, dv_amount]:
        entry.add_data_validation(validation)

    dv_month.add(f"A5:A{ENTRY_END_ROW}")
    dv_year.add(f"B5:B{ENTRY_END_ROW}")
    dv_type.add(f"C5:C{ENTRY_END_ROW}")
    dv_category.add(f"D5:D{ENTRY_END_ROW}")
    dv_amount.add(f"E5:E{ENTRY_END_ROW}")

    entry["K4"] = "Current Category"
    style_cell(entry["K4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    for row in range(5, ENTRY_END_ROW + 1):
        entry[f"K{row}"] = (
            '=IF(D{0}="","",IFERROR(INDEX(ADD_CATEGORY!$A$5:$A$300,'
            'MATCH(D{0},ADD_CATEGORY!$A$5:$A$300,0)),'
            'IFERROR(INDEX(_SYSTEM!$G$5:$G$300,MATCH(D{0},_SYSTEM!$H$5:$H$300,0)),D{0})))'
        ).format(row)
    entry.column_dimensions["K"].hidden = True

    dv_summary_year = DataValidation(type="list", formula1="=FilterYearList", allow_blank=False)
    summary.add_data_validation(dv_summary_year)
    dv_summary_year.add("E5")
    summary["D4"] = "Choose year here"
    summary["D5"] = "Year"
    summary["D6"] = "Click E5 and select the year"
    style_cell(summary["D4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    style_cell(summary["D5"], COLORS["gold_soft"], bold=True)
    style_cell(summary["E5"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    summary["E4"] = "Select"
    style_cell(summary["E4"], COLORS["gold_soft"], bold=True, align="center")
    summary.merge_cells("D6:G6")
    style_cell(summary["D6"], COLORS["cream"])

    start_row = 23
    end_row = start_row + len(ordered_categories) - 1
    for idx in range(start_row, end_row + 1):
        list_row = idx - 18
        summary[f"A{idx}"] = f'=ADD_CATEGORY!$A${list_row}'
        summary[f"C{idx}"] = f'=SUMPRODUCT((ENTRY!$E$5:$E${ENTRY_END_ROW})*--(ENTRY!$K$5:$K${ENTRY_END_ROW}=$A{idx})*--(ENTRY!$C$5:$C${ENTRY_END_ROW}="Receipt"))'
        summary[f"D{idx}"] = f'=SUMPRODUCT((ENTRY!$E$5:$E${ENTRY_END_ROW})*--(ENTRY!$K$5:$K${ENTRY_END_ROW}=$A{idx})*--(ENTRY!$C$5:$C${ENTRY_END_ROW}="Expense"))'
        summary[f"B{idx}"] = f'=IF(AND(C{idx}>0,D{idx}>0),"Both",IF(C{idx}>0,"Receipt only",IF(D{idx}>0,"Expense only","No records")))'
        summary[f"E{idx}"] = f'=N(C{idx})-N(D{idx})'
        summary[f"F{idx}"] = f'=SUMPRODUCT(--(ENTRY!$K$5:$K${ENTRY_END_ROW}=$A{idx})*--(ENTRY!$E$5:$E${ENTRY_END_ROW}>0))'

    for cell_ref, formula in {
        "B5": "=FilterTypeList",
        "B6": "=FilterCategoryList",
        "B7": "=FilterYearList",
        "B8": "=FilterMonthList",
    }.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        details.add_data_validation(dv)
        dv.add(cell_ref)
    details["A4"] = "Choose filters here"
    details["D4"] = "Results"
    details.merge_cells("A9:E9")
    details["A9"] = "Click B5 to B8 and choose values from the dropdown."
    style_cell(details["A4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    style_cell(details["D4"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    style_cell(details["A9"], COLORS["cream"])
    for ref in ["B5", "B6", "B7", "B8"]:
        style_cell(details[ref], COLORS["green"], bold=True, color=COLORS["white"], align="center")

    details["E5"] = (
        f'=SUMPRODUCT((ENTRY!$E$5:$E${ENTRY_END_ROW})*'
        f'--(((ENTRY!$C$5:$C${ENTRY_END_ROW}=$B$5)+($B$5="All"))>0)*'
        f'--(((ENTRY!$K$5:$K${ENTRY_END_ROW}=$B$6)+($B$6="All"))>0)*'
        f'--(((ENTRY!$B$5:$B${ENTRY_END_ROW}=$B$7)+($B$7="All"))>0)*'
        f'--(((ENTRY!$A$5:$A${ENTRY_END_ROW}=$B$8)+($B$8="All"))>0))'
    )
    details["E6"] = (
        f'=SUMPRODUCT(--(ENTRY!$E$5:$E${ENTRY_END_ROW}>0)*'
        f'--(((ENTRY!$C$5:$C${ENTRY_END_ROW}=$B$5)+($B$5="All"))>0)*'
        f'--(((ENTRY!$K$5:$K${ENTRY_END_ROW}=$B$6)+($B$6="All"))>0)*'
        f'--(((ENTRY!$B$5:$B${ENTRY_END_ROW}=$B$7)+($B$7="All"))>0)*'
        f'--(((ENTRY!$A$5:$A${ENTRY_END_ROW}=$B$8)+($B$8="All"))>0))'
    )
    details["E8"] = (
        f'=IF($B$6="All","Mixed",IF(COUNTIFS(ENTRY!$K$5:$K${ENTRY_END_ROW},$B$6,ENTRY!$C$5:$C${ENTRY_END_ROW},"Receipt")>0,'
        f'IF(COUNTIFS(ENTRY!$K$5:$K${ENTRY_END_ROW},$B$6,ENTRY!$C$5:$C${ENTRY_END_ROW},"Expense")>0,"Both","Receipt only"),'
        f'IF(COUNTIFS(ENTRY!$K$5:$K${ENTRY_END_ROW},$B$6,ENTRY!$C$5:$C${ENTRY_END_ROW},"Expense")>0,"Expense only","No records")))'
    )
    details["A11"] = "Matching records"
    details["A12"] = "Entry ID"
    details["B12"] = "Month"
    details["C12"] = "Year"
    details["D12"] = "Type"
    details["E12"] = "Category"
    details["F12"] = "Amount"
    style_cell(details["A11"], None, bold=True, color=COLORS["green"])
    for ref in ["A12", "B12", "C12", "D12", "E12", "F12"]:
        style_cell(details[ref], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    for row in range(13, max(260, ENTRY_END_ROW // 2)):
        details[f"G{row}"] = (
            '=IFERROR(AGGREGATE(15,6,(ROW(ENTRY!$I$5:$I${0})-ROW(ENTRY!$I$5)+1)/('
            '(((ENTRY!$C$5:$C${0}=$B$5)+($B$5="All"))>0)*'
            '(((ENTRY!$K$5:$K${0}=$B$6)+($B$6="All"))>0)*'
            '(((ENTRY!$B$5:$B${0}=$B$7)+($B$7="All"))>0)*'
            '(((ENTRY!$A$5:$A${0}=$B$8)+($B$8="All"))>0)*'
            '(ENTRY!$E$5:$E${0}<>"")),ROWS($G$13:G{1})),"")'
        ).format(ENTRY_END_ROW, row)
        details[f"A{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$I$5:$I${1},$G{0}))'.format(row, ENTRY_END_ROW)
        details[f"B{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$A$5:$A${1},$G{0}))'.format(row, ENTRY_END_ROW)
        details[f"C{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$B$5:$B${1},$G{0}))'.format(row, ENTRY_END_ROW)
        details[f"D{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$C$5:$C${1},$G{0}))'.format(row, ENTRY_END_ROW)
        details[f"E{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$K$5:$K${1},$G{0}))'.format(row, ENTRY_END_ROW)
        details[f"F{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$E$5:$E${1},$G{0}))'.format(row, ENTRY_END_ROW)
        for ref in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}"]:
            style_cell(details[ref])
        details[f"F{row}"].number_format = '"Rs. " #,##0.00'
    details.column_dimensions["G"].hidden = True

    entry.conditional_formatting.add(
        f"D5:D{ENTRY_END_ROW}",
        FormulaRule(
            formula=['AND($D5<>"",COUNTIF(EntryCategoryList,$D5)=0)'],
            fill=PatternFill("solid", fgColor=COLORS["red_soft"]),
        ),
    )

    entry["J3"] = "New category"
    entry["J4"] = "If a category is missing or needs a new name, do it once in ADD_CATEGORY."
    style_cell(entry["J3"], COLORS["green"], bold=True, color=COLORS["white"], align="center")
    style_cell(entry["J4"], COLORS["gold_soft"])
    entry.column_dimensions["J"].width = 30

    if "ADD_CATEGORY" in workbook.sheetnames:
        order = ["README", "ENTRY", "ADD_CATEGORY", "SUMMARY", "DETAILS", "_SYSTEM"]
        workbook._sheets.sort(key=lambda sheet: order.index(sheet.title))

    workbook.save(OUTPUT_PATH)

    print(f"Created {OUTPUT_PATH}")
    print("Normalized rows:")
    for name, count in sorted(normalized_counts.items()):
        print(f"{name}: {count}")
    print(f"Final category count: {len(ordered_categories)}")


if __name__ == "__main__":
    main()
