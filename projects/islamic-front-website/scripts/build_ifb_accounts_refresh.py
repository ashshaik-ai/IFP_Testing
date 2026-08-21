from __future__ import annotations

from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.styles.colors import Color
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


SOURCE_PATH = Path(r"C:\Users\User\OneDrive\Documents\IFB_accounts_clean_master_metrics.xlsx")
OUTPUT_PATH = SOURCE_PATH.with_name("IFB_accounts_clean_master_metrics_refreshed.xlsx")
PROTECT_PASSWORD = "ifb2026"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_TO_NUM = {name: idx for idx, name in enumerate(MONTHS, start=1)}

COLORS = {
    "green": "0F5132",
    "green_alt": "1B5E20",
    "green_soft": "E8F3EC",
    "gold": "C59D2A",
    "gold_soft": "F6EFCF",
    "cream": "F8F5ED",
    "ink": "163029",
    "slate": "5C6C66",
    "border": "CBD8D0",
    "red_soft": "FCE8E6",
    "red": "C62828",
    "white": "FFFFFF",
}

ENTRY_RANGE = {
    "month": "ENTRY!$A$5:$A$5000",
    "year": "ENTRY!$B$5:$B$5000",
    "type": "ENTRY!$C$5:$C$5000",
    "category": "ENTRY!$D$5:$D$5000",
    "amount": "ENTRY!$E$5:$E$5000",
    "month_label": "ENTRY!$F$5:$F$5000",
    "month_start": "ENTRY!$G$5:$G$5000",
    "fy": "ENTRY!$H$5:$H$5000",
}


def load_existing_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["TRANSACTIONS"]
    rows: list[dict[str, object]] = []
    for row_idx in range(3, sheet.max_row + 1):
        month_start = sheet[f"B{row_idx}"].value
        txn_type = sheet[f"E{row_idx}"].value
        category = sheet[f"F{row_idx}"].value
        amount = sheet[f"H{row_idx}"].value
        if not (month_start and txn_type and category and amount not in (None, "")):
            continue
        if not isinstance(month_start, datetime):
            continue
        rows.append(
            {
                "month": month_start.strftime("%b"),
                "year": month_start.year,
                "type": txn_type,
                "category": category,
                "amount": float(amount),
            }
        )
    return rows


def derive_fy(year: int, month_num: int) -> str:
    fy_start = year if month_num >= 8 else year - 1
    return f"{fy_start}-{str(fy_start + 1)[-2:]}"


def auto_fit_columns(sheet, min_width: int = 11, max_width: int = 28) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        longest = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            longest = max(longest, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max(min_width, min(max_width, longest + 2))


def set_named_range(workbook: Workbook, name: str, attr_text: str) -> None:
    workbook.defined_names.pop(name, None)
    workbook.defined_names.add(DefinedName(name, attr_text=attr_text))


def apply_base_sheet_style(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"


def fill_range(sheet, cell_range: str, fill: PatternFill) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = copy(fill)


def add_title_block(sheet, title: str, subtitle: str, end_col: int = 10) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    sheet["A1"] = title
    sheet["A1"].font = Font(size=18, bold=True, color=COLORS["white"])
    sheet["A1"].fill = PatternFill("solid", fgColor=COLORS["green"])
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, italic=True, color=COLORS["ink"])
    sheet["A2"].fill = PatternFill("solid", fgColor=COLORS["cream"])
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 34


def header_style(cell, fill_color: str) -> None:
    cell.font = Font(bold=True, color=COLORS["white"])
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        bottom=Side(style="thin", color=COLORS["border"]),
    )


def body_style(cell, fill_color: str | None = None, align: str = "left") -> None:
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.border = Border(
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        bottom=Side(style="thin", color=COLORS["border"]),
    )
    cell.alignment = Alignment(horizontal=align, vertical="center")


def style_currency_column(sheet, column: str, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        sheet[f"{column}{row}"].number_format = '"Rs. " #,##0.00'


def style_int_column(sheet, column: str, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        sheet[f"{column}{row}"].number_format = "#,##0"


def style_percent_column(sheet, column: str, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        sheet[f"{column}{row}"].number_format = "0.0%"


def enable_data_labels(chart) -> None:
    labels = DataLabelList()
    labels.showVal = True
    labels.showLegendKey = False
    labels.showCatName = False
    chart.dLbls = labels


def make_lists_sheet(workbook: Workbook, categories: list[str], years: list[int], fys: list[str]) -> None:
    sheet = workbook.create_sheet("LISTS")
    add_title_block(
        sheet,
        "Anjuman Records Lists",
        "Managed by Islamic Front Board. This sheet stores dropdown values used by the other sheets.",
        end_col=18,
    )
    apply_base_sheet_style(sheet)

    headers = {
        "A4": "Months",
        "B4": "Month No",
        "D4": "Types",
        "G4": "Categories",
        "J4": "Years",
        "M4": "FY",
        "O4": "Type Filter",
        "P4": "Category Filter",
        "Q4": "FY Filter",
        "R4": "Year Filter",
        "S4": "Month Filter",
    }
    for cell_ref, text in headers.items():
        sheet[cell_ref] = text
        header_style(sheet[cell_ref], COLORS["green_alt"])

    for idx, month in enumerate(MONTHS, start=5):
        sheet[f"A{idx}"] = month
        sheet[f"B{idx}"] = MONTH_TO_NUM[month]
    for row in range(5, 17):
        body_style(sheet[f"A{row}"], COLORS["green_soft"])
        body_style(sheet[f"B{row}"], COLORS["cream"], align="center")

    types = ["Receipt", "Expense"]
    for idx, txn_type in enumerate(types, start=5):
        sheet[f"D{idx}"] = txn_type
        body_style(sheet[f"D{idx}"], COLORS["gold_soft"])

    for idx, category in enumerate(categories, start=5):
        sheet[f"G{idx}"] = category
        body_style(sheet[f"G{idx}"])

    for idx, year in enumerate(years, start=5):
        sheet[f"J{idx}"] = year
        body_style(sheet[f"J{idx}"], COLORS["green_soft"], align="center")

    for idx, fy in enumerate(fys, start=5):
        sheet[f"M{idx}"] = fy
        body_style(sheet[f"M{idx}"], COLORS["gold_soft"], align="center")

    type_filter = ["All", *types]
    category_filter = ["All", *categories]
    fy_filter = ["All", *fys]
    year_filter = ["All", *years]
    month_filter = ["All", *MONTHS]
    filter_columns = {
        "O": type_filter,
        "P": category_filter,
        "Q": fy_filter,
        "R": year_filter,
        "S": month_filter,
    }
    for column, values in filter_columns.items():
        for idx, value in enumerate(values, start=5):
            sheet[f"{column}{idx}"] = value
            body_style(sheet[f"{column}{idx}"], COLORS["cream"])

    auto_fit_columns(sheet)


def make_entry_sheet(
    workbook: Workbook, rows: list[dict[str, object]], categories: list[str], years: list[int], fys: list[str]
) -> None:
    sheet = workbook.create_sheet("ENTRY")
    add_title_block(
        sheet,
        "Anjuman Monthly Records Entry",
        "Managed by Islamic Front Board. Staff only need to enter Month, Year, Type, Category, and Amount. Everything else updates automatically.",
        end_col=20,
    )
    apply_base_sheet_style(sheet)
    sheet["A3"] = "Staff should fill only these columns"
    sheet["A3"].font = Font(bold=True, color=COLORS["green"])
    sheet["B3"] = "Month"
    sheet["C3"] = "Year"
    sheet["D3"] = "Type"
    sheet["E3"] = "Category"
    sheet["F3"] = "Amount"
    for ref in ["B3", "C3", "D3", "E3", "F3"]:
        body_style(sheet[ref], COLORS["gold_soft"], align="center")

    headers = ["Month", "Year", "Type", "Category", "Amount", "Month Label", "Month Start", "FY", "Entry ID"]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col_idx, value=header)
        header_style(cell, COLORS["green"])

    start_row = 5
    for offset, row in enumerate(rows):
        current = start_row + offset
        sheet[f"A{current}"] = row["month"]
        sheet[f"B{current}"] = row["year"]
        sheet[f"C{current}"] = row["type"]
        sheet[f"D{current}"] = row["category"]
        sheet[f"E{current}"] = row["amount"]
        sheet[f"F{current}"] = '=IF(OR(A{0}="",B{0}=""),"",TEXT(G{0},"mmm yyyy"))'.format(current)
        sheet[f"G{current}"] = '=IF(OR(A{0}="",B{0}=""),"",DATE(B{0},MATCH(A{0},$K$5:$K$16,0),1))'.format(current)
        sheet[f"H{current}"] = (
            '=IF(G{0}="","",TEXT(DATE(YEAR(G{0})-(MONTH(G{0})<8),8,1),"yyyy")&"-"&'
            'TEXT(DATE(YEAR(G{0})-(MONTH(G{0})<8)+1,8,1),"yy"))'
        ).format(current)
        sheet[f"I{current}"] = '=IF(COUNTA(A{0}:E{0})=0,"","TXN-"&TEXT(ROW()-4,"0000"))'.format(current)

    for row in range(start_row, start_row + max(len(rows), 1) + 350):
        for col in range(1, 10):
            body_style(sheet.cell(row=row, column=col), COLORS["white"] if col <= 5 else COLORS["cream"])

    table_end = start_row + len(rows) - 1
    table = Table(displayName="EntryTable", ref=f"A4:I{table_end}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for row in range(start_row, table_end + 1):
        sheet[f"E{row}"].number_format = '"Rs. " #,##0.00'
        sheet[f"B{row}"].number_format = "0"
        sheet[f"G{row}"].number_format = "mmm yyyy"

    # Hidden helper and dropdown source columns so the workbook can stay at two sheets only.
    sheet["K4"] = "Months"
    sheet["L4"] = "Types"
    sheet["M4"] = "Categories"
    sheet["N4"] = "Years"
    sheet["O4"] = "FY"
    sheet["P4"] = "Type Filter"
    sheet["Q4"] = "Category Filter"
    sheet["R4"] = "FY Filter"
    sheet["S4"] = "Year Filter"
    sheet["T4"] = "Month Filter"

    types = ["Receipt", "Expense"]
    for idx, month in enumerate(MONTHS, start=5):
        sheet[f"K{idx}"] = month
    for idx, txn_type in enumerate(types, start=5):
        sheet[f"L{idx}"] = txn_type
    for idx, category in enumerate(categories, start=5):
        sheet[f"M{idx}"] = category
    for idx, year in enumerate(years, start=5):
        sheet[f"N{idx}"] = year
    for idx, fy in enumerate(fys, start=5):
        sheet[f"O{idx}"] = fy

    for idx, value in enumerate(["All", *types], start=5):
        sheet[f"P{idx}"] = value
    for idx, value in enumerate(["All", *categories], start=5):
        sheet[f"Q{idx}"] = value
    for idx, value in enumerate(["All", *fys], start=5):
        sheet[f"R{idx}"] = value
    for idx, value in enumerate(["All", *years], start=5):
        sheet[f"S{idx}"] = value
    for idx, value in enumerate(["All", *MONTHS], start=5):
        sheet[f"T{idx}"] = value

    for column in ["F", "G", "H", "I", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        sheet.column_dimensions[column].hidden = True

    month_end = 4 + len(MONTHS)
    year_end = 4 + len(years)
    category_end = 4 + len(categories)
    fy_end = 4 + len(fys)

    set_named_range(workbook, "EntryMonthList", f"ENTRY!$K$5:$K${month_end}")
    set_named_range(workbook, "EntryTypeList", "ENTRY!$L$5:$L$6")
    set_named_range(workbook, "EntryCategoryList", f"ENTRY!$M$5:$M${category_end}")
    set_named_range(workbook, "EntryYearList", f"ENTRY!$N$5:$N${year_end}")
    set_named_range(workbook, "EntryFyList", f"ENTRY!$O$5:$O${fy_end}")
    set_named_range(workbook, "FilterTypeList", "ENTRY!$P$5:$P$7")
    set_named_range(workbook, "FilterCategoryList", f"ENTRY!$Q$5:$Q${5 + len(categories)}")
    set_named_range(workbook, "FilterFyList", f"ENTRY!$R$5:$R${5 + len(fys)}")
    set_named_range(workbook, "FilterYearList", f"ENTRY!$S$5:$S${5 + len(years)}")
    set_named_range(workbook, "FilterMonthList", "ENTRY!$T$5:$T$17")

    dv_month = DataValidation(type="list", formula1="=EntryMonthList", allow_blank=True)
    dv_year = DataValidation(type="list", formula1="=EntryYearList", allow_blank=True)
    dv_type = DataValidation(type="list", formula1="=EntryTypeList", allow_blank=True)
    dv_category = DataValidation(type="list", formula1="=EntryCategoryList", allow_blank=True)
    dv_amount = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=True)
    for validation in [dv_month, dv_year, dv_type, dv_category, dv_amount]:
        sheet.add_data_validation(validation)
    editable_end = start_row + 5000
    dv_month.add(f"A{start_row}:A{editable_end}")
    dv_year.add(f"B{start_row}:B{editable_end}")
    dv_type.add(f"C{start_row}:C{editable_end}")
    dv_category.add(f"D{start_row}:D{editable_end}")
    dv_amount.add(f"E{start_row}:E{editable_end}")

    for row in range(start_row, editable_end + 1):
        for column in ["A", "B", "C", "D", "E"]:
            sheet[f"{column}{row}"].fill = PatternFill("solid", fgColor=COLORS["green_soft"])
            sheet[f"{column}{row}"].protection = Protection(locked=False)
        for column in ["F", "G", "H", "I"]:
            sheet[f"{column}{row}"].protection = Protection(locked=True)

    sheet.conditional_formatting.add(
        f"E{start_row}:E{editable_end}",
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["red_soft"])),
    )
    auto_fit_columns(sheet)
    sheet.column_dimensions["D"].width = 34


def build_month_rows(start_date: datetime, years_forward: int) -> list[datetime]:
    rows = []
    year = start_date.year
    month = start_date.month
    end_year = start_date.year + years_forward
    end_month = 12
    current = datetime(year, month, 1)
    while current <= datetime(end_year, end_month, 1):
        rows.append(current)
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    return rows


def make_summary_sheet(workbook: Workbook, categories: list[str], latest_year: int) -> None:
    sheet = workbook.create_sheet("SUMMARY")
    add_title_block(
        sheet,
        "Anjuman Summary",
        "Managed by Islamic Front Board. This page shows total money, month-wise totals for a chosen year, and category usage.",
        end_col=10,
    )
    apply_base_sheet_style(sheet)
    sheet.freeze_panes = "A4"

    sheet["A4"] = "Overall totals"
    header_style(sheet["A4"], COLORS["green"])
    sheet["D4"] = "Month-wise table"
    header_style(sheet["D4"], COLORS["green"])
    overall_rows = [
        (5, "Money received", f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*({ENTRY_RANGE["type"]}="Receipt"))'),
        (6, "Money spent", f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*({ENTRY_RANGE["type"]}="Expense"))'),
        (7, "Balance", "=B5-B6"),
        (8, "Number of records", f'=SUMPRODUCT(--({ENTRY_RANGE["amount"]}>0))'),
        (9, "Latest year with data", latest_year),
    ]
    for row, label, formula in overall_rows:
        sheet[f"A{row}"] = label
        body_style(sheet[f"A{row}"], COLORS["gold_soft"])
        sheet[f"B{row}"] = formula
        body_style(sheet[f"B{row}"], COLORS["green_soft"], align="center")

    for ref in ["B5", "B6", "B7"]:
        sheet[ref].number_format = '"Rs. " #,##0.00'
    sheet["B8"].number_format = "#,##0"

    sheet["D5"] = "Year"
    body_style(sheet["D5"], COLORS["gold_soft"])
    sheet["E5"] = latest_year
    body_style(sheet["E5"], COLORS["green_soft"], align="center")
    dv = DataValidation(type="list", formula1="=FilterYearList", allow_blank=False)
    sheet.add_data_validation(dv)
    dv.add("E5")

    sheet["D6"] = "Meaning"
    body_style(sheet["D6"], COLORS["gold_soft"])
    sheet["E6"] = "Balance = Money received - Money spent"
    body_style(sheet["E6"], COLORS["cream"])

    month_headers = ["Month", "Money Received", "Money Spent", "Balance"]
    for idx, header in enumerate(month_headers, start=4):
        header_style(sheet.cell(row=8, column=idx, value=header), COLORS["green_alt"])
    for idx, month in enumerate(MONTHS, start=9):
        sheet[f"D{idx}"] = month
        sheet[f"E{idx}"] = (
            f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*--({ENTRY_RANGE["type"]}="Receipt")*--({ENTRY_RANGE["month"]}=$D{idx})*--({ENTRY_RANGE["year"]}=$E$5))'
        )
        sheet[f"F{idx}"] = (
            f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*--({ENTRY_RANGE["type"]}="Expense")*--({ENTRY_RANGE["month"]}=$D{idx})*--({ENTRY_RANGE["year"]}=$E$5))'
        )
        sheet[f"G{idx}"] = f'=N(E{idx})-N(F{idx})'
        for col in range(4, 8):
            body_style(sheet.cell(row=idx, column=col))
    style_currency_column(sheet, "E", 9, 20)
    style_currency_column(sheet, "F", 9, 20)
    style_currency_column(sheet, "G", 9, 20)

    start_row = 23
    end_row = start_row + len(categories) - 1
    sheet["A21"] = "Category usage and totals"
    sheet["A21"].font = Font(bold=True, color=COLORS["green"])
    category_headers = ["Category", "Used as", "Money Received", "Money Spent", "Balance", "Record Count"]
    for idx, header in enumerate(category_headers, start=1):
        header_style(sheet.cell(row=22, column=idx, value=header), COLORS["green_alt"])

    for idx in range(start_row, end_row + 1):
        list_row = idx - 18
        sheet[f"A{idx}"] = f'=ENTRY!$M${list_row}'
        sheet[f"C{idx}"] = f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*--({ENTRY_RANGE["category"]}=$A{idx})*--({ENTRY_RANGE["type"]}="Receipt"))'
        sheet[f"D{idx}"] = f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*--({ENTRY_RANGE["category"]}=$A{idx})*--({ENTRY_RANGE["type"]}="Expense"))'
        sheet[f"B{idx}"] = (
            f'=IF(AND(C{idx}>0,D{idx}>0),"Both",IF(C{idx}>0,"Receipt only",IF(D{idx}>0,"Expense only","No records")))'
        )
        sheet[f"E{idx}"] = f'=N(C{idx})-N(D{idx})'
        sheet[f"F{idx}"] = f'=SUMPRODUCT(--({ENTRY_RANGE["category"]}=$A{idx})*--({ENTRY_RANGE["amount"]}>0))'
        for col in range(1, 7):
            body_style(sheet.cell(row=idx, column=col))
    style_currency_column(sheet, "C", start_row, end_row)
    style_currency_column(sheet, "D", start_row, end_row)
    style_currency_column(sheet, "E", start_row, end_row)
    style_int_column(sheet, "F", start_row, end_row)

    sheet.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["red_soft"])),
    )
    auto_fit_columns(sheet)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 18
    sheet.column_dimensions["G"].width = 18


def make_details_sheet(workbook: Workbook, rows: list[dict[str, object]], categories: list[str], latest_year: int) -> None:
    sheet = workbook.create_sheet("DETAILS")
    add_title_block(
        sheet,
        "Anjuman Category Details",
        "Managed by Islamic Front Board. Choose Type, Category, Year, and Month. Use Type = Receipt to see revenue.",
        end_col=9,
    )
    apply_base_sheet_style(sheet)
    sheet.freeze_panes = "A12"

    sheet["A4"] = "Filters"
    header_style(sheet["A4"], COLORS["green"])
    filter_rows = [
        (5, "Type", "B5", "All"),
        (6, "Category", "B6", "All"),
        (7, "Year", "B7", latest_year),
        (8, "Month", "B8", "All"),
    ]
    for row, label, ref, default in filter_rows:
        sheet[f"A{row}"] = label
        body_style(sheet[f"A{row}"], COLORS["gold_soft"])
        sheet[ref] = default
        body_style(sheet[ref], COLORS["green_soft"], align="center")

    validations = {
        "B5": "=FilterTypeList",
        "B6": "=FilterCategoryList",
        "B7": "=FilterYearList",
        "B8": "=FilterMonthList",
    }
    for cell_ref, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        sheet.add_data_validation(dv)
        dv.add(cell_ref)

    sheet["D4"] = "Answers"
    header_style(sheet["D4"], COLORS["green"])
    answer_rows = [
        (5, "Total amount"),
        (6, "Number of records"),
        (7, "Average amount"),
        (8, "Category status"),
    ]
    for row, label in answer_rows:
        sheet[f"D{row}"] = label
        body_style(sheet[f"D{row}"], COLORS["cream"])
        body_style(sheet[f"E{row}"], COLORS["green_soft"], align="center")

    sheet["E5"] = (
        f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*'
        f'--((({ENTRY_RANGE["type"]}=$B$5)+($B$5="All"))>0)*'
        f'--((({ENTRY_RANGE["category"]}=$B$6)+($B$6="All"))>0)*'
        f'--((({ENTRY_RANGE["year"]}=$B$7)+($B$7="All"))>0)*'
        f'--((({ENTRY_RANGE["month"]}=$B$8)+($B$8="All"))>0))'
    )
    sheet["E6"] = (
        f'=SUMPRODUCT(--({ENTRY_RANGE["amount"]}>0)*'
        f'--((({ENTRY_RANGE["type"]}=$B$5)+($B$5="All"))>0)*'
        f'--((({ENTRY_RANGE["category"]}=$B$6)+($B$6="All"))>0)*'
        f'--((({ENTRY_RANGE["year"]}=$B$7)+($B$7="All"))>0)*'
        f'--((({ENTRY_RANGE["month"]}=$B$8)+($B$8="All"))>0))'
    )
    sheet["E7"] = '=IFERROR(E5/E6,0)'
    sheet["D8"] = "Category use"
    body_style(sheet["D8"], COLORS["cream"])
    sheet["E8"] = (
        '=IF($B$6="All","Mixed",IF(COUNTIFS(ENTRY!$D$5:$D$5000,$B$6,ENTRY!$C$5:$C$5000,"Receipt")>0,'
        'IF(COUNTIFS(ENTRY!$D$5:$D$5000,$B$6,ENTRY!$C$5:$C$5000,"Expense")>0,"Both","Receipt only"),'
        'IF(COUNTIFS(ENTRY!$D$5:$D$5000,$B$6,ENTRY!$C$5:$C$5000,"Expense")>0,"Expense only","No records")))'
    )
    sheet["E5"].number_format = sheet["E7"].number_format = '"Rs. " #,##0.00'
    sheet["E6"].number_format = "#,##0"

    sheet["A11"] = "Matching records"
    sheet["A11"].font = Font(bold=True, color=COLORS["green"])
    headers = ["Entry ID", "Month", "Year", "Type", "Category", "Amount"]
    for idx, header in enumerate(headers, start=1):
        header_style(sheet.cell(row=12, column=idx, value=header), COLORS["green_alt"])
    output_start = 13
    output_end = max(output_start + len(rows) + 25, 260)
    for row in range(output_start, output_end + 1):
        sheet[f"G{row}"] = (
            '=IFERROR(AGGREGATE(15,6,(ROW(ENTRY!$I$5:$I$5000)-ROW(ENTRY!$I$5)+1)/('
            '(((ENTRY!$C$5:$C$5000=$B$5)+($B$5="All"))>0)*'
            '(((ENTRY!$D$5:$D$5000=$B$6)+($B$6="All"))>0)*'
            '(((ENTRY!$B$5:$B$5000=$B$7)+($B$7="All"))>0)*'
            '(((ENTRY!$A$5:$A$5000=$B$8)+($B$8="All"))>0)*'
            '(ENTRY!$E$5:$E$5000<>"")),ROWS($G$13:G{0})),"")'
        ).format(row)
        sheet[f"A{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$I$5:$I$5000,$G{0}))'.format(row)
        sheet[f"B{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$A$5:$A$5000,$G{0}))'.format(row)
        sheet[f"C{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$B$5:$B$5000,$G{0}))'.format(row)
        sheet[f"D{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$C$5:$C$5000,$G{0}))'.format(row)
        sheet[f"E{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$D$5:$D$5000,$G{0}))'.format(row)
        sheet[f"F{row}"] = '=IF($G{0}="","",INDEX(ENTRY!$E$5:$E$5000,$G{0}))'.format(row)
        for col in range(1, 8):
            body_style(sheet.cell(row=row, column=col))
        sheet[f"F{row}"].number_format = '"Rs. " #,##0.00'

    auto_fit_columns(sheet)
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["E"].width = 30
    sheet.column_dimensions["G"].hidden = True


def make_summary_monthly_sheet(workbook: Workbook, month_rows: list[datetime]) -> None:
    sheet = workbook.create_sheet("SUMMARY_MONTHLY")
    add_title_block(
        sheet,
        "Monthly Summary",
        "Managed by Islamic Front Board. This sheet shows money received, money spent, and balance month by month.",
        end_col=8,
    )
    apply_base_sheet_style(sheet)
    headers = ["Month Start", "Month", "Financial Year", "Money Received", "Money Spent", "Balance", "Receipt Records", "Expense Records"]
    for col_idx, header in enumerate(headers, start=1):
        header_style(sheet.cell(row=4, column=col_idx, value=header), COLORS["green_alt"])

    for idx, month_start in enumerate(month_rows, start=5):
        sheet[f"A{idx}"] = month_start
        sheet[f"A{idx}"].number_format = "mmm yyyy"
        sheet[f"B{idx}"] = f'=TEXT(A{idx},"mmm yyyy")'
        sheet[f"C{idx}"] = (
            f'=TEXT(DATE(YEAR(A{idx})-(MONTH(A{idx})<8),8,1),"yyyy")&"-"&'
            f'TEXT(DATE(YEAR(A{idx})-(MONTH(A{idx})<8)+1,8,1),"yy")'
        )
        sheet[f"D{idx}"] = (
            f'=IF(COUNTIFS({ENTRY_RANGE["month_start"]},$A{idx},{ENTRY_RANGE["type"]},"Receipt")=0,"",'
            f'SUMIFS({ENTRY_RANGE["amount"]},{ENTRY_RANGE["month_start"]},$A{idx},{ENTRY_RANGE["type"]},"Receipt"))'
        )
        sheet[f"E{idx}"] = (
            f'=IF(COUNTIFS({ENTRY_RANGE["month_start"]},$A{idx},{ENTRY_RANGE["type"]},"Expense")=0,"",'
            f'SUMIFS({ENTRY_RANGE["amount"]},{ENTRY_RANGE["month_start"]},$A{idx},{ENTRY_RANGE["type"]},"Expense"))'
        )
        sheet[f"F{idx}"] = f'=IF(AND(D{idx}="",E{idx}=""),"",N(D{idx})-N(E{idx}))'
        sheet[f"G{idx}"] = f'=COUNTIFS({ENTRY_RANGE["month_start"]},$A{idx},{ENTRY_RANGE["type"]},"Receipt",{ENTRY_RANGE["amount"]},">0")'
        sheet[f"H{idx}"] = f'=COUNTIFS({ENTRY_RANGE["month_start"]},$A{idx},{ENTRY_RANGE["type"]},"Expense",{ENTRY_RANGE["amount"]},">0")'
        for col in range(1, 9):
            body_style(sheet.cell(row=idx, column=col))

    style_currency_column(sheet, "D", 5, 5 + len(month_rows) - 1)
    style_currency_column(sheet, "E", 5, 5 + len(month_rows) - 1)
    style_currency_column(sheet, "F", 5, 5 + len(month_rows) - 1)
    style_int_column(sheet, "G", 5, 5 + len(month_rows) - 1)
    style_int_column(sheet, "H", 5, 5 + len(month_rows) - 1)
    sheet.conditional_formatting.add(
        f"F5:F{4 + len(month_rows)}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["red_soft"])),
    )
    auto_fit_columns(sheet)


def make_summary_category_sheet(workbook: Workbook, category_count: int = 200) -> None:
    sheet = workbook.create_sheet("SUMMARY_CATEGORY")
    add_title_block(
        sheet,
        "Category Summary",
        "Managed by Islamic Front Board. This sheet shows money received, money spent, and balance for each category.",
        end_col=6,
    )
    apply_base_sheet_style(sheet)
    headers = ["Category", "Money Received", "Money Spent", "Balance", "Record Count", "Average Record"]
    for col_idx, header in enumerate(headers, start=1):
        header_style(sheet.cell(row=4, column=col_idx, value=header), COLORS["green_alt"])

    for idx in range(5, 5 + category_count):
        list_row = idx
        sheet[f"A{idx}"] = f'=IF(LISTS!$G${list_row}="","",LISTS!$G${list_row})'
        sheet[f"B{idx}"] = f'=IF($A{idx}="","",SUMIFS({ENTRY_RANGE["amount"]},{ENTRY_RANGE["category"]},$A{idx},{ENTRY_RANGE["type"]},"Receipt"))'
        sheet[f"C{idx}"] = f'=IF($A{idx}="","",SUMIFS({ENTRY_RANGE["amount"]},{ENTRY_RANGE["category"]},$A{idx},{ENTRY_RANGE["type"]},"Expense"))'
        sheet[f"D{idx}"] = f'=IF($A{idx}="","",N(B{idx})-N(C{idx}))'
        sheet[f"E{idx}"] = f'=IF($A{idx}="","",COUNTIFS({ENTRY_RANGE["category"]},$A{idx},{ENTRY_RANGE["amount"]},">0"))'
        sheet[f"F{idx}"] = f'=IFERROR((N(B{idx})+N(C{idx}))/E{idx},"")'
        for col in range(1, 7):
            body_style(sheet.cell(row=idx, column=col))

    style_currency_column(sheet, "B", 5, 4 + category_count)
    style_currency_column(sheet, "C", 5, 4 + category_count)
    style_currency_column(sheet, "D", 5, 4 + category_count)
    style_currency_column(sheet, "F", 5, 4 + category_count)
    style_int_column(sheet, "E", 5, 4 + category_count)
    sheet.conditional_formatting.add(
        f"D5:D{4 + category_count}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["red_soft"])),
    )
    auto_fit_columns(sheet)
    sheet.column_dimensions["A"].width = 36


def make_metrics_sheet(workbook: Workbook, years: list[int], fy_values: list[str], category_count: int = 200) -> None:
    sheet = workbook.create_sheet("METRICS")
    add_title_block(
        sheet,
        "Anjuman Questions and Metrics",
        "Managed by Islamic Front Board. Choose filters, read the answers, then see one simple bar graph at the end.",
        end_col=14,
    )
    apply_base_sheet_style(sheet)
    sheet.freeze_panes = "A4"

    sheet.merge_cells("A4:B4")
    sheet["A4"] = "Step 1: Choose filters"
    header_style(sheet["A4"], COLORS["green"])
    sheet.merge_cells("D4:E4")
    sheet["D4"] = "Step 2: Read the answer"
    header_style(sheet["D4"], COLORS["green"])
    sheet.merge_cells("G4:H4")
    sheet["G4"] = "Year-wise amount"
    header_style(sheet["G4"], COLORS["green"])

    filter_rows = [
        (5, "Type"),
        (6, "Category"),
        (7, "Financial Year"),
        (8, "Calendar Year"),
        (9, "Month"),
    ]
    for row, label in filter_rows:
        sheet[f"A{row}"] = label
        body_style(sheet[f"A{row}"], COLORS["gold_soft"])

    filter_defaults = {"B5": "All", "B6": "All", "B7": "All", "B8": "All", "B9": "All"}
    for cell_ref, value in filter_defaults.items():
        sheet[cell_ref] = value
        body_style(sheet[cell_ref], COLORS["green_soft"], align="center")
        sheet[cell_ref].protection = Protection(locked=False)

    validations = {
        "B5": "=LISTS!$O$5:$O$7",
        "B6": "=LISTS!$P$5:$P$200",
        "B7": "=LISTS!$Q$5:$Q$30",
        "B8": "=LISTS!$R$5:$R$30",
        "B9": "=LISTS!$S$5:$S$17",
    }
    for cell_ref, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        sheet.add_data_validation(dv)
        dv.add(cell_ref)

    sheet["D5"] = "Question"
    sheet["E5"] = "Answer"
    header_style(sheet["D5"], COLORS["green_alt"])
    header_style(sheet["E5"], COLORS["green_alt"])
    answer_rows = [
        (6, "Total amount"),
        (7, "Number of records"),
        (8, "Average amount"),
        (9, "Selected category"),
    ]
    for row, label in answer_rows:
        sheet[f"D{row}"] = label
        body_style(sheet[f"D{row}"], COLORS["cream"])
        body_style(sheet[f"E{row}"], COLORS["green_soft"], align="center")

    sheet["E6"] = (
        f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*'
        f'--((({ENTRY_RANGE["type"]}=$B$5)+($B$5="All"))>0)*'
        f'--((({ENTRY_RANGE["category"]}=$B$6)+($B$6="All"))>0)*'
        f'--((({ENTRY_RANGE["fy"]}=$B$7)+($B$7="All"))>0)*'
        f'--((({ENTRY_RANGE["year"]}=$B$8)+($B$8="All"))>0)*'
        f'--((({ENTRY_RANGE["month"]}=$B$9)+($B$9="All"))>0))'
    )
    sheet["E7"] = (
        f'=SUMPRODUCT(--({ENTRY_RANGE["amount"]}>0)*'
        f'--((({ENTRY_RANGE["type"]}=$B$5)+($B$5="All"))>0)*'
        f'--((({ENTRY_RANGE["category"]}=$B$6)+($B$6="All"))>0)*'
        f'--((({ENTRY_RANGE["fy"]}=$B$7)+($B$7="All"))>0)*'
        f'--((({ENTRY_RANGE["year"]}=$B$8)+($B$8="All"))>0)*'
        f'--((({ENTRY_RANGE["month"]}=$B$9)+($B$9="All"))>0))'
    )
    sheet["E8"] = '=IFERROR(E6/E7,0)'
    sheet["E9"] = '=IF($B$6="All","All categories",$B$6)'
    sheet["E6"].number_format = sheet["E8"].number_format = '"Rs. " #,##0.00'
    sheet["E7"].number_format = "#,##0"

    sheet["G5"] = "Year"
    sheet["H5"] = "Amount"
    header_style(sheet["G5"], COLORS["green_alt"])
    header_style(sheet["H5"], COLORS["green_alt"])
    for idx, year in enumerate(years, start=6):
        sheet[f"G{idx}"] = year
        sheet[f"H{idx}"] = (
            f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*--((({ENTRY_RANGE["type"]}=$B$5)+($B$5="All"))>0)*'
            f'--((({ENTRY_RANGE["category"]}=$B$6)+($B$6="All"))>0)*--({ENTRY_RANGE["year"]}=$G{idx}))'
        )
        body_style(sheet[f"G{idx}"])
        body_style(sheet[f"H{idx}"])
    style_currency_column(sheet, "H", 6, 5 + len(years))

    sheet["A12"] = "Category summary for the current filter"
    sheet["A12"].font = Font(bold=True, color=COLORS["green"])
    table_headers = ["Category", "Money Received", "Money Spent", "Balance", "Record Count", "Balance Share"]
    for idx, header in enumerate(table_headers, start=1):
        header_style(sheet.cell(row=13, column=idx, value=header), COLORS["green_alt"])

    start_row = 14
    end_row = start_row + category_count - 1
    for idx in range(start_row, end_row + 1):
        list_row = idx - 9
        sheet[f"A{idx}"] = f'=IF(LISTS!$G${list_row}="","",LISTS!$G${list_row})'
        sheet[f"B{idx}"] = (
            f'=IF($A{idx}="","",SUMPRODUCT(({ENTRY_RANGE["amount"]})*--({ENTRY_RANGE["category"]}=$A{idx})*--({ENTRY_RANGE["type"]}="Receipt")*'
            f'--((({ENTRY_RANGE["fy"]}=$B$7)+($B$7="All"))>0)*--((({ENTRY_RANGE["year"]}=$B$8)+($B$8="All"))>0)*'
            f'--((({ENTRY_RANGE["month"]}=$B$9)+($B$9="All"))>0)))'
        )
        sheet[f"C{idx}"] = (
            f'=IF($A{idx}="","",SUMPRODUCT(({ENTRY_RANGE["amount"]})*--({ENTRY_RANGE["category"]}=$A{idx})*--({ENTRY_RANGE["type"]}="Expense")*'
            f'--((({ENTRY_RANGE["fy"]}=$B$7)+($B$7="All"))>0)*--((({ENTRY_RANGE["year"]}=$B$8)+($B$8="All"))>0)*'
            f'--((({ENTRY_RANGE["month"]}=$B$9)+($B$9="All"))>0)))'
        )
        sheet[f"D{idx}"] = f'=IF($A{idx}="","",N(B{idx})-N(C{idx}))'
        sheet[f"E{idx}"] = (
            f'=IF($A{idx}="","",SUMPRODUCT(--({ENTRY_RANGE["category"]}=$A{idx})*--({ENTRY_RANGE["amount"]}>0)*'
            f'--((({ENTRY_RANGE["fy"]}=$B$7)+($B$7="All"))>0)*--((({ENTRY_RANGE["year"]}=$B$8)+($B$8="All"))>0)*'
            f'--((({ENTRY_RANGE["month"]}=$B$9)+($B$9="All"))>0)))'
        )
        sheet[f"F{idx}"] = f'=IFERROR(D{idx}/SUM($D${start_row}:$D${end_row}),0)'
        for col in range(1, 7):
            body_style(sheet.cell(row=idx, column=col))

    style_currency_column(sheet, "B", start_row, end_row)
    style_currency_column(sheet, "C", start_row, end_row)
    style_currency_column(sheet, "D", start_row, end_row)
    style_int_column(sheet, "E", start_row, end_row)
    style_percent_column(sheet, "F", start_row, end_row)

    year_chart = BarChart()
    year_chart.type = "col"
    year_chart.title = "Selected amount by year"
    year_chart.y_axis.title = "Rupees"
    year_chart.x_axis.title = "Calendar Year"
    year_chart.height = 8
    year_chart.width = 12
    year_chart.style = 10
    data = Reference(sheet, min_col=8, min_row=5, max_row=5 + len(years))
    cats = Reference(sheet, min_col=7, min_row=6, max_row=5 + len(years))
    year_chart.add_data(data, titles_from_data=True)
    year_chart.set_categories(cats)
    enable_data_labels(year_chart)
    chart_row = end_row + 4
    sheet.add_chart(year_chart, f"A{chart_row}")

    sheet.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["red_soft"])),
    )
    auto_fit_columns(sheet)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 16


def make_dashboard_sheet(workbook: Workbook, fy_values: list[str], default_fy: str) -> None:
    sheet = workbook.create_sheet("DASHBOARD")
    add_title_block(
        sheet,
        "Anjuman Records Dashboard",
        "Managed by Islamic Front Board. This page gives a short summary only. No charts here.",
        end_col=8,
    )
    apply_base_sheet_style(sheet)
    sheet.freeze_panes = "A4"

    sheet["A4"] = "Big picture summary"
    header_style(sheet["A4"], COLORS["green"])
    sheet["D4"] = "Monthly summary"
    header_style(sheet["D4"], COLORS["green"])

    summary_rows = [
        (5, "Money received", f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*({ENTRY_RANGE["type"]}="Receipt"))'),
        (6, "Money spent", f'=SUMPRODUCT(({ENTRY_RANGE["amount"]})*({ENTRY_RANGE["type"]}="Expense"))'),
        (7, "Balance", "=B5-B6"),
        (8, "Number of records", f'=SUMPRODUCT(--({ENTRY_RANGE["amount"]}>0))'),
        (9, "Highest receipt category", '=INDEX(SUMMARY_CATEGORY!$A$5:$A$204,MATCH(MAX(SUMMARY_CATEGORY!$B$5:$B$204),SUMMARY_CATEGORY!$B$5:$B$204,0))'),
        (10, "Highest expense category", '=INDEX(SUMMARY_CATEGORY!$A$5:$A$204,MATCH(MAX(SUMMARY_CATEGORY!$C$5:$C$204),SUMMARY_CATEGORY!$C$5:$C$204,0))'),
    ]
    for row, label, formula in summary_rows:
        sheet[f"A{row}"] = label
        body_style(sheet[f"A{row}"], COLORS["gold_soft"])
        sheet[f"B{row}"] = formula
        body_style(sheet[f"B{row}"], COLORS["green_soft"], align="center")

    for ref in ["B5", "B6", "B7"]:
        sheet[ref].number_format = '"Rs. " #,##0.00'
    sheet["B8"].number_format = "#,##0"

    sheet["D5"] = "Financial year"
    body_style(sheet["D5"], COLORS["gold_soft"])
    sheet["E5"] = default_fy if default_fy else (fy_values[-1] if fy_values else "All")
    body_style(sheet["E5"], COLORS["green_soft"], align="center")
    dv = DataValidation(type="list", formula1="=LISTS!$Q$5:$Q$30", allow_blank=False)
    sheet.add_data_validation(dv)
    dv.add("E5")
    sheet["E5"].protection = Protection(locked=False)

    sheet["D6"] = "Meaning"
    body_style(sheet["D6"], COLORS["gold_soft"])
    sheet["E6"] = "Received - Spent = Balance"
    body_style(sheet["E6"], COLORS["cream"])

    headers = ["Month", "Money Received", "Money Spent", "Balance"]
    for idx, header in enumerate(headers, start=4):
        header_style(sheet.cell(row=8, column=idx, value=header), COLORS["green_alt"])
    for idx in range(9, 21):
        month_number = ((idx - 9 + 7) % 12) + 1
        month_name = MONTHS[month_number - 1]
        sheet[f"D{idx}"] = month_name
        sheet[f"E{idx}"] = (
            f'=SUMIFS(SUMMARY_MONTHLY!$D:$D,SUMMARY_MONTHLY!$C:$C,$E$5,SUMMARY_MONTHLY!$B:$B,$D{idx}&" "&LEFT($E$5,4))+'
            f'SUMIFS(SUMMARY_MONTHLY!$D:$D,SUMMARY_MONTHLY!$C:$C,$E$5,SUMMARY_MONTHLY!$B:$B,$D{idx}&" "&("20"&RIGHT($E$5,2)))'
        )
        sheet[f"F{idx}"] = (
            f'=SUMIFS(SUMMARY_MONTHLY!$E:$E,SUMMARY_MONTHLY!$C:$C,$E$5,SUMMARY_MONTHLY!$B:$B,$D{idx}&" "&LEFT($E$5,4))+'
            f'SUMIFS(SUMMARY_MONTHLY!$E:$E,SUMMARY_MONTHLY!$C:$C,$E$5,SUMMARY_MONTHLY!$B:$B,$D{idx}&" "&("20"&RIGHT($E$5,2)))'
        )
        sheet[f"G{idx}"] = f'=N(E{idx})-N(F{idx})'
        for col in range(4, 8):
            body_style(sheet.cell(row=idx, column=col))

    style_currency_column(sheet, "E", 9, 20)
    style_currency_column(sheet, "F", 9, 20)
    style_currency_column(sheet, "G", 9, 20)

    sheet.merge_cells("A12:G12")
    sheet["A12"] = "Use the ENTRY sheet for data entry and the METRICS sheet for detailed questions."
    body_style(sheet["A12"], COLORS["cream"])
    sheet["A12"].alignment = Alignment(wrap_text=True)

    auto_fit_columns(sheet)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 18
    sheet.column_dimensions["G"].width = 18

def reorder_sheets(workbook: Workbook) -> None:
    order = ["ENTRY", "SUMMARY", "DETAILS"]
    workbook._sheets.sort(key=lambda sheet: order.index(sheet.title))
    workbook.active = 0


def main() -> None:
    rows = load_existing_rows(SOURCE_PATH)
    if not rows:
        raise SystemExit("No transaction rows found in the source workbook.")

    categories = sorted({str(row["category"]) for row in rows})
    min_year = min(int(row["year"]) for row in rows)
    max_year = max(int(row["year"]) for row in rows)
    current_year = datetime.now().year
    year_end = max(max_year + 8, current_year + 5)
    years = list(range(min_year, year_end + 1))
    fy_values = [derive_fy(year, 8) for year in range(min_year, year_end + 1)]

    month_starts = [datetime(int(row["year"]), MONTH_TO_NUM[str(row["month"])], 1) for row in rows]
    first_month = min(month_starts)
    latest_month = max(month_starts)
    latest_year = latest_month.year
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    make_entry_sheet(workbook, rows, categories, years, fy_values)
    make_summary_sheet(workbook, categories, latest_year)
    make_details_sheet(workbook, rows, categories, latest_year)
    reorder_sheets(workbook)

    candidates = [
        OUTPUT_PATH,
        SOURCE_PATH.with_name("IFB_accounts_clean_master_metrics_refreshed_v2.xlsx"),
        SOURCE_PATH.with_name("IFB_accounts_clean_master_metrics_refreshed_v3.xlsx"),
        SOURCE_PATH.with_name("IFB_accounts_clean_master_metrics_refreshed_v4.xlsx"),
        SOURCE_PATH.with_name("IFB_accounts_clean_master_metrics_refreshed_v5.xlsx"),
        SOURCE_PATH.with_name("IFB_accounts_clean_master_metrics_refreshed_v6.xlsx"),
    ]
    final_output = None
    last_error = None
    for candidate in candidates:
        try:
            workbook.save(candidate)
            final_output = candidate
            break
        except PermissionError as exc:
            last_error = exc
    if final_output is None:
        raise last_error
    print(f"Created {final_output}")


if __name__ == "__main__":
    main()
