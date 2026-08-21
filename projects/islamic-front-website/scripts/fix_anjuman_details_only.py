from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


WORKBOOK_PATH = Path(r"C:\Users\User\OneDrive\Documents 1\Anjuman_IFB_accounts.xlsx")
BACKUP_PATH = Path(r"C:\Users\User\OneDrive\Documents 1\Anjuman_IFB_accounts_backup_before_details_fix.xlsx")

COLORS = {
    "green": "0F5132",
    "cream": "F8F5ED",
    "border": "CBD8D0",
    "white": "FFFFFF",
}


def set_named_range(workbook, name: str, attr_text: str) -> None:
    workbook.defined_names.pop(name, None)
    workbook.defined_names.add(DefinedName(name, attr_text=attr_text))


def clear_details_validations(details) -> None:
    details.data_validations.dataValidation = []


def thin_border() -> Border:
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def style_cell(cell, fill: str | None = None, bold: bool = False, color: str = "000000", align: str = "left") -> None:
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.font = Font(bold=bold, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def last_entry_row(entry) -> int:
    for row in range(entry.max_row, 4, -1):
        if any(entry[f"{col}{row}"].value not in (None, "") for col in ["A", "B", "C", "D", "E", "I"]):
            return row
    return 4


def allowed_types_for_category(entry, end_row: int, category: str) -> set[str]:
    types: set[str] = set()
    if category in (None, "", "All"):
        return {"All", "Receipt", "Expense"}
    for row in range(5, end_row + 1):
        row_category = entry[f"D{row}"].value
        row_type = entry[f"C{row}"].value
        if row_category in (None, "") or row_type in (None, ""):
            continue
        if str(row_category).strip() == category:
            types.add(str(row_type).strip())
    if not types:
        return {"All"}
    return {"All", *types}


def main() -> None:
    shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)

    workbook = load_workbook(WORKBOOK_PATH)
    entry = workbook["ENTRY"]
    details = workbook["DETAILS"]
    end_row = last_entry_row(entry)

    try:
        workbook.calculation.forceFullCalc = True
        workbook.calculation.fullCalcOnLoad = True
    except AttributeError:
        pass

    details["J4"] = "All Types"
    details["J5"] = "All"
    details["J6"] = "Receipt"
    details["J7"] = "Expense"
    details["K4"] = "Receipt Types"
    details["K5"] = "All"
    details["K6"] = "Receipt"
    details["L4"] = "Expense Types"
    details["L5"] = "All"
    details["L6"] = "Expense"
    details["M4"] = "Only All"
    details["M5"] = "All"
    for col in ["J", "K", "L", "M"]:
        details.column_dimensions[col].hidden = True

    set_named_range(workbook, "DetailsTypeAll", "DETAILS!$J$5:$J$7")
    set_named_range(workbook, "DetailsTypeReceiptOnly", "DETAILS!$K$5:$K$6")
    set_named_range(workbook, "DetailsTypeExpenseOnly", "DETAILS!$L$5:$L$6")
    set_named_range(workbook, "DetailsTypeAllOnly", "DETAILS!$M$5:$M$5")

    clear_details_validations(details)
    validations = {
        "B5": '=INDIRECT(IF($E$8="Receipt only","DetailsTypeReceiptOnly",IF($E$8="Expense only","DetailsTypeExpenseOnly",IF($E$8="No records","DetailsTypeAllOnly","DetailsTypeAll"))))',
        "B6": "=FilterCategoryList",
        "B7": "=FilterYearList",
        "B8": "=FilterMonthList",
    }
    for cell_ref, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        details.add_data_validation(dv)
        dv.add(cell_ref)

    current_category = "" if details["B6"].value is None else str(details["B6"].value).strip()
    current_type = "" if details["B5"].value is None else str(details["B5"].value).strip()
    allowed_types = allowed_types_for_category(entry, end_row, current_category)
    if current_type not in allowed_types:
        details["B5"] = "All"

    details["A9"] = "Choose Category first. Type will show only allowed options."
    details.merge_cells("A9:F9")
    style_cell(details["A9"], COLORS["cream"])

    details["A11"] = "Matching records"
    style_cell(details["A11"], None, bold=True, color=COLORS["green"])

    headers = {
        "A12": "Entry ID",
        "B12": "Month",
        "C12": "Year",
        "D12": "Type",
        "E12": "Category",
        "F12": "Amount",
    }
    for ref, value in headers.items():
        details[ref] = value
        style_cell(details[ref], COLORS["green"], bold=True, color=COLORS["white"], align="center")

    entry["AA4"] = 0
    for row in range(5, end_row + 1):
        entry[f"AA{row}"] = (
            '=IF(AND($E{0}<>"",'
            '((($C{0}=DETAILS!$B$5)+(DETAILS!$B$5="All"))>0),'
            '((($D{0}=DETAILS!$B$6)+(DETAILS!$B$6="All"))>0),'
            '((($B{0}=DETAILS!$B$7)+(DETAILS!$B$7="All"))>0),'
            '((($A{0}=DETAILS!$B$8)+(DETAILS!$B$8="All"))>0)),'
            'MAX($AA$4:AA{1})+1,"")'
        ).format(row, row - 1)
    entry.column_dimensions["AA"].hidden = True

    clear_to_row = max(details.max_row, end_row + 20)
    for row in range(13, clear_to_row + 1):
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            details[f"{col}{row}"] = None
        details[f"F{row}"].number_format = '"Rs. " #,##0.00'

    for row in range(13, clear_to_row + 1):
        nth = row - 12
        details[f"A{row}"] = f'=IFERROR(INDEX(ENTRY!$I$5:$I${end_row},MATCH({nth},ENTRY!$AA$5:$AA${end_row},0)),"")'
        details[f"B{row}"] = f'=IF($A{row}="","",INDEX(ENTRY!$A$5:$A${end_row},MATCH({nth},ENTRY!$AA$5:$AA${end_row},0)))'
        details[f"C{row}"] = f'=IF($A{row}="","",INDEX(ENTRY!$B$5:$B${end_row},MATCH({nth},ENTRY!$AA$5:$AA${end_row},0)))'
        details[f"D{row}"] = f'=IF($A{row}="","",INDEX(ENTRY!$C$5:$C${end_row},MATCH({nth},ENTRY!$AA$5:$AA${end_row},0)))'
        details[f"E{row}"] = f'=IF($A{row}="","",INDEX(ENTRY!$D$5:$D${end_row},MATCH({nth},ENTRY!$AA$5:$AA${end_row},0)))'
        details[f"F{row}"] = f'=IF($A{row}="","",INDEX(ENTRY!$E$5:$E${end_row},MATCH({nth},ENTRY!$AA$5:$AA${end_row},0)))'
        for col in ["A", "B", "C", "D", "E", "F"]:
            style_cell(details[f"{col}{row}"])
        details[f"F{row}"].number_format = '"Rs. " #,##0.00'

    workbook.save(WORKBOOK_PATH)
    print(f"Backup created: {BACKUP_PATH}")
    print(f"Updated workbook: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
