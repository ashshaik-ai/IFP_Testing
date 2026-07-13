from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.storage import load_jobs, read_json, voters_path, write_json

_SERIAL_HEADERS = {"serial", "serial no", "serial no.", "serial_no", "serial number", "sl no", "sl.no", "s.no", "sno"}
_WHATSAPP_HEADERS = {"whatsapp", "whatsapp?", "has whatsapp", "wa", "wa status"}
_YES_VALUES = {"yes", "y", "true", "1"}
_NO_VALUES = {"no", "n", "false", "0"}


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_columns(header_row: tuple) -> tuple[int, int]:
    serial_col = wa_col = -1
    for idx, cell in enumerate(header_row):
        label = _cell_str(cell).lower()
        if serial_col == -1 and label in _SERIAL_HEADERS:
            serial_col = idx
        if wa_col == -1 and label in _WHATSAPP_HEADERS:
            wa_col = idx
    return serial_col, wa_col


def import_whatsapp_status_from_xlsx(file_bytes: bytes) -> dict:
    """Read a (Serial No, WhatsApp) sheet -- WhatsApp column holding
    Yes/No -- and set `has_whatsapp` on matching voters, matched by
    `serial_no` across every job. Blank/unrecognized WhatsApp cells are
    skipped (row is ignored, not treated as No). Serial numbers are
    globally unique across this deployment's voter lists."""
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {"total_rows": 0, "status_rows": 0, "matched": 0, "updated": 0,
                 "marked_yes": 0, "marked_no": 0, "not_found_count": 0}

    serial_col, wa_col = _find_columns(header)
    if serial_col == -1 or wa_col == -1:
        serial_col, wa_col = 0, 2

    status_by_serial: dict[str, bool] = {}
    total_rows = 0
    for row in rows:
        if row is None or all(cell is None for cell in row):
            continue
        total_rows += 1
        serial = _cell_str(row[serial_col]) if serial_col < len(row) else ""
        raw_status = _cell_str(row[wa_col]).lower() if wa_col < len(row) else ""
        if not serial or not raw_status:
            continue
        if raw_status in _YES_VALUES:
            status_by_serial[serial] = True
        elif raw_status in _NO_VALUES:
            status_by_serial[serial] = False

    matched_serials: set[str] = set()
    updated = marked_yes = marked_no = 0
    for job in load_jobs():
        path = voters_path(job["id"])
        voters = read_json(path, [])
        changed = False
        for voter in voters:
            serial = _cell_str(voter.get("serial_no", ""))
            if serial not in status_by_serial:
                continue
            matched_serials.add(serial)
            status = status_by_serial[serial]
            if voter.get("has_whatsapp") != status:
                voter["has_whatsapp"] = status
                changed = True
                updated += 1
                if status:
                    marked_yes += 1
                else:
                    marked_no += 1
            # Default opt-in rule: a mobile number + a confirmed-WhatsApp
            # status together imply opt-in; anything else does not. Recomputed
            # every time has_whatsapp is (re)confirmed by this importer, even
            # if has_whatsapp itself didn't change value this run.
            desired_optin = bool(str(voter.get("mobile", "")).strip()) and status is True
            if bool(voter.get("wa_optin")) != desired_optin:
                voter["wa_optin"] = desired_optin
                changed = True
        if changed:
            write_json(path, voters)

    not_found_count = len(set(status_by_serial) - matched_serials)
    return {
        "total_rows": total_rows,
        "status_rows": len(status_by_serial),
        "matched": len(matched_serials),
        "updated": updated,
        "marked_yes": marked_yes,
        "marked_no": marked_no,
        "not_found_count": not_found_count,
    }
