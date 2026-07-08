from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.storage import load_jobs, read_json, voters_path, write_json

_SERIAL_HEADERS = {"serial", "serial no", "serial no.", "serial_no", "serial number", "sl no", "sl.no", "s.no", "sno"}


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_serial_column(header_row: tuple) -> int:
    for idx, cell in enumerate(header_row):
        if _cell_str(cell).lower() in _SERIAL_HEADERS:
            return idx
    return -1


def import_flags_from_xlsx(file_bytes: bytes) -> dict:
    """Read a sheet of serial numbers (one per row) and flag matching voters
    (`is_flagged = True`), matched by `serial_no` across every job. Already
    flagged voters are left untouched. Serial numbers are globally unique
    across this deployment's voter lists, so no per-job scoping is needed."""
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {"total_rows": 0, "serial_rows": 0, "matched": 0, "flagged": 0,
                "already_flagged": 0, "not_found_count": 0}

    serial_col = _find_serial_column(header)
    if serial_col == -1:
        serial_col = 0

    serials: set[str] = set()
    total_rows = 0
    for row in rows:
        if row is None or all(cell is None for cell in row):
            continue
        total_rows += 1
        serial = _cell_str(row[serial_col]) if serial_col < len(row) else ""
        if serial:
            serials.add(serial)

    matched_serials: set[str] = set()
    flagged = 0
    already_flagged = 0
    for job in load_jobs():
        path = voters_path(job["id"])
        voters = read_json(path, [])
        changed = False
        for voter in voters:
            serial = _cell_str(voter.get("serial_no", ""))
            if serial not in serials:
                continue
            matched_serials.add(serial)
            if voter.get("is_flagged"):
                already_flagged += 1
                continue
            voter["is_flagged"] = True
            changed = True
            flagged += 1
        if changed:
            write_json(path, voters)

    not_found_count = len(serials - matched_serials)
    return {
        "total_rows": total_rows,
        "serial_rows": len(serials),
        "matched": len(matched_serials),
        "flagged": flagged,
        "already_flagged": already_flagged,
        "not_found_count": not_found_count,
    }
