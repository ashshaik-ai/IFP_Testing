from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.storage import load_jobs, read_json, voters_path, write_json

_SERIAL_HEADERS = {"serial", "serial no", "serial no.", "serial_no", "serial number", "sl no", "sl.no", "s.no", "sno"}
_PHONE_HEADERS = {"phone", "phone no", "phone no.", "phone number", "mobile", "mobile no", "mobile number", "ph.no", "ph no", "contact", "contact no"}


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_phone(value: object) -> str:
    """Only a plain 10-digit local mobile number is accepted — anything else
    (missing digits, a country code, letters) is treated as no phone."""
    digits = "".join(ch for ch in _cell_str(value) if ch.isdigit())
    return digits if len(digits) == 10 else ""


def _find_columns(header_row: tuple) -> tuple[int, int]:
    serial_col = phone_col = -1
    for idx, cell in enumerate(header_row):
        label = _cell_str(cell).lower()
        if serial_col == -1 and label in _SERIAL_HEADERS:
            serial_col = idx
        if phone_col == -1 and label in _PHONE_HEADERS:
            phone_col = idx
    return serial_col, phone_col


def import_phones_from_xlsx(file_bytes: bytes) -> dict:
    """Read a (Serial No, Phone No) sheet and write matching phone numbers
    onto voters' `mobile` field, matched by `serial_no` across every job.
    Serial numbers are globally unique across this deployment's voter lists
    (contiguous ranges per source PDF), so no per-job scoping is needed."""
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {"total_rows": 0, "phone_rows": 0, "matched": 0, "updated": 0, "not_found_count": 0}

    serial_col, phone_col = _find_columns(header)
    if serial_col == -1 or phone_col == -1:
        serial_col, phone_col = 0, 1

    phone_by_serial: dict[str, str] = {}
    total_rows = 0
    for row in rows:
        if row is None or all(cell is None for cell in row):
            continue
        total_rows += 1
        serial = _cell_str(row[serial_col]) if serial_col < len(row) else ""
        phone = _norm_phone(row[phone_col]) if phone_col < len(row) else ""
        if serial and phone:
            phone_by_serial[serial] = phone

    matched_serials: set[str] = set()
    updated = 0
    for job in load_jobs():
        path = voters_path(job["id"])
        voters = read_json(path, [])
        changed = False
        for voter in voters:
            phone = phone_by_serial.get(_cell_str(voter.get("serial_no", "")))
            if not phone:
                continue
            matched_serials.add(_cell_str(voter.get("serial_no", "")))
            if voter.get("mobile") != phone:
                voter["mobile"] = phone
                changed = True
                updated += 1
        if changed:
            write_json(path, voters)

    not_found_count = len(set(phone_by_serial) - matched_serials)
    return {
        "total_rows": total_rows,
        "phone_rows": len(phone_by_serial),
        "matched": len(matched_serials),
        "updated": updated,
        "not_found_count": not_found_count,
    }
